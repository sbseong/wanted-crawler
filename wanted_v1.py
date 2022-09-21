link_num = int(input("몇번째 채용공고부터 시작할지 숫자로 입력해주세요.(예> 200) >>>"))

import os
import time
import requests
from selenium import webdriver
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from datetime import datetime
import pyautogui
import sys



# 크롬 드라이버 자동 업데이트

from webdriver_manager.chrome import ChromeDriverManager

pyautogui.FAILSAFE = False


# 상대경로 세팅
if getattr(sys, 'frozen', False):
    #test.exe로 실행한 경우,test.exe를 보관한 디렉토리의 full path를 취득
    program_directory = os.path.dirname(os.path.abspath(sys.executable))
else:
    #python test.py로 실행한 경우,test.py를 보관한 디렉토리의 full path를 취득
    program_directory = os.path.dirname(os.path.abspath(__file__))

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

err_log_path = program_directory + "/Job_Error_log.txt"
save_date = datetime.today().strftime("%Y%m%d_%H%M")
exl_name = program_directory + f"/job_{save_date}.xlsx"
exl_sample_name = program_directory + f"/job_crawling_sample.xlsx"




# 브라우저 꺼짐 방지
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

# 불필요한 에러 메시지 삭제
chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)
driver.implicitly_wait(5) # 웹페이지가 로딩될때까지 5초 기다림
# driver.maximize_window() # 화면 최대화

# 크롤링 방지 설정을 undefined로 변경
driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
    "source": """
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            })
            """
})


# 엑셀 세팅


wb = openpyxl.load_workbook(exl_sample_name)
ws = wb.active




gonggo_num = 1

# 메인 페이지 이동
wait = WebDriverWait(driver, 5)
url = "https://www.wanted.co.kr/wdlist/518?country=kr&job_sort=company.response_rate_order&years=0&years=2&selected=655&selected=899&selected=1025&locations=all" #1
driver.get(url=url)
time.sleep(5)




# link_num = 1
exl_num = 1


while True:

    ids = driver.find_elements(by = By.CSS_SELECTOR, value = ".Card_className__u5rsb")
    time.sleep(2)

    for id in ids :

        if gonggo_num < link_num :
            gonggo_num += 1

        if gonggo_num >= link_num :
            try :
                link = ids[link_num-1].find_element(By.CSS_SELECTOR, f"#__next > div.JobList_cn__t_THp > div > div > div.List_List_container__JnQMS > ul > li:nth-child({link_num}) > div > a").get_attribute('href')
                co_name = ids[link_num-1].find_element(By.CSS_SELECTOR, f"#__next > div.JobList_cn__t_THp > div > div > div.List_List_container__JnQMS > ul > li:nth-child({link_num}) > div > a > div > div.job-card-company-name").text
                print(f"{link_num} : {co_name}")
                time.sleep(1)
                link_num += 1
                gonggo_num += 1

                # pyautogui.hotkey('ctrl', 't') # 윈도우용
                pyautogui.hotkey("command", "t") # 맥용
                time.sleep(2)  
                all_windows = driver.window_handles
                driver.switch_to.window(all_windows[1])
                
                driver.get(link)
                time.sleep(3)


                # 무한 스크롤
                while True:

                    driver.find_element(By.CSS_SELECTOR, "body").send_keys(Keys.END)


                    try : 
                        locate2 = driver.find_element(By.CSS_SELECTOR, "#__next > div.JobDetail_cn__WezJh > div.JobDetail_contentWrapper__DQDB6 > div.JobDetail_relativeWrapper__F9DT5 > div.JobContent_className___ca57 > div.JobContent_descriptionWrapper__SM4UD > section.JobWorkPlace_className__ra6rp > div:nth-child(2) > span.header").text
                        time.sleep(1)

                        break
                    except :
                        time.sleep(1)
                        pass

                # 항목별 데이터 수집            

                try :
                    jikgun = driver.find_element(By.CSS_SELECTOR, "#__next > div.JobDetail_cn__WezJh > div.JobDetail_contentWrapper__DQDB6 > div.JobDetail_relativeWrapper__F9DT5 > div.JobContent_className___ca57 > section.JobHeader_className__HttDA > h2").text
                except:
                    jikgun = ""


                try :
                    closing_date = driver.find_element(By.CSS_SELECTOR, "#__next > div.JobDetail_cn__WezJh > div.JobDetail_contentWrapper__DQDB6 > div.JobDetail_relativeWrapper__F9DT5 > div.JobContent_className___ca57 > div.JobContent_descriptionWrapper__SM4UD > section.JobWorkPlace_className__ra6rp > div:nth-child(1) > span.body").text
                except:
                    closing_date = ""

                try :
                    locate = driver.find_element(By.CSS_SELECTOR, "#__next > div.JobDetail_cn__WezJh > div.JobDetail_contentWrapper__DQDB6 > div.JobDetail_relativeWrapper__F9DT5 > div.JobContent_className___ca57 > div.JobContent_descriptionWrapper__SM4UD > section.JobWorkPlace_className__ra6rp > div:nth-child(2) > span.body").text
                except:
                    locate = ""


                jikmu = ""
                upmu = ""
                jakyuk = ""
                woodea = ""
                bokri = ""
                skill = ""

                i = 1

                while i < 11 :
                    try :
                        el_name = driver.find_element(By.CSS_SELECTOR, f"#__next > div.JobDetail_cn__WezJh > div.JobDetail_contentWrapper__DQDB6 > div.JobDetail_relativeWrapper__F9DT5 > div.JobContent_className___ca57 > div.JobContent_descriptionWrapper__SM4UD > section.JobDescription_JobDescription__VWfcb > h6:nth-child({i})").text

                        if el_name == "주요업무" :

                            upmu2 = driver.find_element(By.CSS_SELECTOR, f"#__next > div.JobDetail_cn__WezJh > div.JobDetail_contentWrapper__DQDB6 > div.JobDetail_relativeWrapper__F9DT5 > div.JobContent_className___ca57 > div.JobContent_descriptionWrapper__SM4UD > section.JobDescription_JobDescription__VWfcb > p:nth-child({i+1})").text
                            upmu = str(upmu2).replace('-','•').replace('ㆍ','•').replace('■','•').replace('●','•').replace('◈','•')

                        if el_name == "자격요건" :

                            jakyuk2 = driver.find_element(By.CSS_SELECTOR, f"#__next > div.JobDetail_cn__WezJh > div.JobDetail_contentWrapper__DQDB6 > div.JobDetail_relativeWrapper__F9DT5 > div.JobContent_className___ca57 > div.JobContent_descriptionWrapper__SM4UD > section.JobDescription_JobDescription__VWfcb > p:nth-child({i+1})").text
                            jakyuk = str(jakyuk2).replace('-','•').replace('ㆍ','•').replace('■','•').replace('●','•').replace('◈','•')
            
                        if el_name == "우대사항" :

                            woodea2 = driver.find_element(By.CSS_SELECTOR, f"#__next > div.JobDetail_cn__WezJh > div.JobDetail_contentWrapper__DQDB6 > div.JobDetail_relativeWrapper__F9DT5 > div.JobContent_className___ca57 > div.JobContent_descriptionWrapper__SM4UD > section.JobDescription_JobDescription__VWfcb > p:nth-child({i+1})").text
                            woodea = str(woodea2).replace('-','•').replace('ㆍ','•').replace('■','•').replace('●','•').replace('◈','•')

                        if el_name == "혜택 및 복지" :

                            bokri2 = driver.find_element(By.CSS_SELECTOR, f"#__next > div.JobDetail_cn__WezJh > div.JobDetail_contentWrapper__DQDB6 > div.JobDetail_relativeWrapper__F9DT5 > div.JobContent_className___ca57 > div.JobContent_descriptionWrapper__SM4UD > section.JobDescription_JobDescription__VWfcb > p:nth-child({i+1})").text
                            bokri = str(bokri2).replace('-','•').replace('ㆍ','•').replace('■','•').replace('●','•').replace('◈','•')

                        if el_name == "기술스택 ・ 툴" :
                            try :
                                jikmu2_1 = driver.find_element(By.CSS_SELECTOR, f"#__next > div.JobDetail_cn__WezJh > div.JobDetail_contentWrapper__DQDB6 > div.JobDetail_relativeWrapper__F9DT5 > div.JobContent_className___ca57 > div.JobContent_descriptionWrapper__SM4UD > section.JobDescription_JobDescription__VWfcb > p:nth-child({i+1}) > div > div:nth-child(1)").text
                            except :
                                jikmu2_1 = ""
                            try :
                                jikmu2_2 = driver.find_element(By.CSS_SELECTOR, f"#__next > div.JobDetail_cn__WezJh > div.JobDetail_contentWrapper__DQDB6 > div.JobDetail_relativeWrapper__F9DT5 > div.JobContent_className___ca57 > div.JobContent_descriptionWrapper__SM4UD > section.JobDescription_JobDescription__VWfcb > p:nth-child({i+1}) > div > div:nth-child(2)").text
                            except :
                                jikmu2_2 = ""
                            try :
                                jikmu2_3 = driver.find_element(By.CSS_SELECTOR, f"#__next > div.JobDetail_cn__WezJh > div.JobDetail_contentWrapper__DQDB6 > div.JobDetail_relativeWrapper__F9DT5 > div.JobContent_className___ca57 > div.JobContent_descriptionWrapper__SM4UD > section.JobDescription_JobDescription__VWfcb > p:nth-child({i+1}) > div > div:nth-child(3)").text
                            except :
                                jikmu2_3 = ""

                            skill = jikmu2_1 + " " + jikmu2_2 + " " + jikmu2_3

                        i += 1

                    except :
                        i += 1
                        pass
                      
                # 엑셀 저장

                ws[f'A{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',jikgun))
                ws[f'B{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',co_name))
                # ws[f'C{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',jikmu))
                ws[f'D{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',upmu))
                ws[f'E{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',jakyuk))
                ws[f'F{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',woodea))
                ws[f'G{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',bokri))
                ws[f'H{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',closing_date))
                ws[f'I{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',locate))
                # ws[f'J{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',contact))
                ws[f'K{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',link))
                ws[f'L{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',skill))

                wb.save(exl_name)
                print("저장완료")
                exl_num += 1
                time.sleep(1)



                driver.close()
                driver.switch_to.window(all_windows[0])

            except :
                break
    
    time.sleep(1)
    driver.find_element(By.CSS_SELECTOR, "body").send_keys(Keys.END)
    time.sleep(3)



while(True):
    pass